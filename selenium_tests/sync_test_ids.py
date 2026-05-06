"""
Cập nhật Test Report với Test ID đúng từ 02_system - Copy.xlsx
và cập nhật 02_system - Copy.xlsx với kết quả test.

Bảng ánh xạ Test ID (selenium → system):
  Registration:
    TC_REG_001    → REG_01  (Đăng ký với dữ liệu hợp lệ)
    TC_REG_002    → REG_02  (Bỏ trống trường bắt buộc)
    TC_REG_005    → REG_03  (Email sai định dạng)
    TC_REG_006    → REG_05  (Mật khẩu yếu - quá ngắn)
    TC_REG_007    → REG_05  (Mật khẩu không có chữ hoa)
    TC_REG_010    → REG_05  (Mật khẩu không có ký tự đặc biệt)
    TC_REG_UI_001 → INT_01  (Kiểm tra giao diện tổng thể)
  Login:
    TC_LOGIN_001     → LOG_01 (Đăng nhập thành công)
    TC_LOGIN_002     → LOG_02 (Sai mật khẩu)
    TC_LOGIN_003     → LOG_02 (Email trống)
    TC_LOGIN_004     → LOG_02 (Mật khẩu trống)
    TC_LOGIN_005     → LOG_03 (Email không tồn tại)
    TC_LOGIN_UI_001  → INT_02 (Giao diện trang đăng nhập)
    TC_LOGIN_UI_002  → LOG_04 (Chức năng quên mật khẩu)
  Income:
    TC_INC_001       → IM_IC_03 (Thêm thu nhập hợp lệ)
    TC_INC_002       → IM_IC_07 (Danh sách cập nhật sau khi thêm)
    TC_INC_005       → IM_IC_05 (Nguồn thu nhập bắt buộc)
    TC_INC_008       → IM_IC_06 (Số tiền phải > 0)
    TC_INC_DELETE_001→ IM_IC_03 (Xóa thu nhập)
    TC_INC_UI_001    → IM_01   (Giao diện trang thu nhập)
  Expense (chi tiêu - không có sheet riêng, dùng tương đương IM):
    TC_EXP_001       → EXP_IC_03 (Thêm chi tiêu hợp lệ)
    TC_EXP_002       → EXP_IC_07 (Danh sách cập nhật sau khi thêm)
    TC_EXP_005       → EXP_IC_05 (Danh mục bắt buộc)
    TC_EXP_008       → EXP_IC_06 (Số tiền phải > 0)
    TC_EXP_DELETE_001→ EXP_IC_03 (Xóa chi tiêu)
    TC_EXP_UI_001    → EXP_01   (Giao diện trang chi tiêu)
  Dashboard & Navigation (không có sheet riêng):
    TC_DASH_001 → DASH_01, TC_DASH_002 → DASH_02, v.v.
    TC_NAV_001  → NAV_01, v.v.
"""
import openpyxl, glob, os, shutil
from datetime import datetime

# ── Bảng ánh xạ ID ──────────────────────────────────────────────────────────
ID_MAP = {
    # Registration
    'TC_REG_001':       ('REG_01',   'Đăng ký với dữ liệu hợp lệ'),
    'TC_REG_002':       ('REG_02',   'Bỏ trống trường bắt buộc - họ tên'),
    'TC_REG_005':       ('REG_03',   'Kiểm tra email sai định dạng'),
    'TC_REG_006':       ('REG_05',   'Kiểm tra mật khẩu yếu - quá ngắn'),
    'TC_REG_007':       ('REG_05',   'Kiểm tra mật khẩu không có chữ hoa'),
    'TC_REG_010':       ('REG_05',   'Kiểm tra mật khẩu không có ký tự đặc biệt'),
    'TC_REG_UI_001':    ('INT_01',   'Kiểm tra giao diện tổng thể trang đăng ký'),
    # Login
    'TC_LOGIN_001':     ('LOG_01',   'Xác minh đăng nhập thành công'),
    'TC_LOGIN_002':     ('LOG_02',   'Xác minh đăng nhập với mật khẩu sai'),
    'TC_LOGIN_003':     ('LOG_02',   'Xác minh đăng nhập với email trống'),
    'TC_LOGIN_004':     ('LOG_02',   'Xác minh đăng nhập với mật khẩu trống'),
    'TC_LOGIN_005':     ('LOG_03',   'Xác minh đăng nhập với email không tồn tại'),
    'TC_LOGIN_UI_001':  ('INT_02',   'Kiểm tra giao diện trang đăng nhập'),
    'TC_LOGIN_UI_002':  ('LOG_04',   'Xác minh chức năng quên mật khẩu'),
    # Income Management
    'TC_INC_001':       ('IM_IC_03', 'Xác minh thêm thu nhập với dữ liệu hợp lệ'),
    'TC_INC_002':       ('IM_IC_07', 'Xác minh danh sách thu nhập cập nhật sau khi thêm'),
    'TC_INC_005':       ('IM_IC_05', 'Xác minh trường nguồn thu nhập là bắt buộc'),
    'TC_INC_008':       ('IM_IC_06', 'Xác minh số tiền phải lớn hơn 0'),
    'TC_INC_DELETE_001':('IM_IC_03', 'Xác minh xóa bản ghi thu nhập'),
    'TC_INC_UI_001':    ('IM_01',    'Kiểm tra giao diện tổng thể trang thu nhập'),
    # Expense (chi tiêu)
    'TC_EXP_001':       ('EXP_IC_03','Xác minh thêm chi tiêu với dữ liệu hợp lệ'),
    'TC_EXP_002':       ('EXP_IC_07','Xác minh danh sách chi tiêu cập nhật sau khi thêm'),
    'TC_EXP_005':       ('EXP_IC_05','Xác minh trường danh mục là bắt buộc'),
    'TC_EXP_008':       ('EXP_IC_06','Xác minh số tiền phải lớn hơn 0'),
    'TC_EXP_DELETE_001':('EXP_IC_03','Xác minh xóa bản ghi chi tiêu'),
    'TC_EXP_UI_001':    ('EXP_01',   'Kiểm tra giao diện tổng thể trang chi tiêu'),
    # Dashboard
    'TC_DASH_001':      ('DASH_01',  'Trang tổng quan tải thành công sau đăng nhập'),
    'TC_DASH_002':      ('DASH_02',  'Trang tổng quan hiển thị thẻ thông tin'),
    'TC_DASH_003':      ('DASH_03',  'Dữ liệu trang tổng quan khớp với API'),
    'TC_DASH_004':      ('DASH_04',  'Điều hướng từ tổng quan sang thu nhập'),
    'TC_DASH_005':      ('DASH_05',  'Điều hướng từ tổng quan sang chi tiêu'),
    'TC_DASH_006':      ('DASH_06',  'Trang tổng quan hiển thị biểu đồ'),
    'TC_DASH_007':      ('DASH_07',  'Trang tổng quan hiển thị giao dịch gần đây'),
    'TC_DASH_UI_001':   ('DASH_UI_01','Kiểm tra bố cục giao diện trang tổng quan'),
    # Navigation
    'TC_NAV_001':       ('NAV_01',   'Chuyển hướng khi chưa đăng nhập'),
    'TC_NAV_002':       ('NAV_02',   'Chuyển hướng trang thu nhập khi chưa đăng nhập'),
    'TC_NAV_003':       ('NAV_03',   'Điều hướng qua thanh sidebar'),
    'TC_NAV_004':       ('NAV_04',   'Điều hướng từ đăng nhập sang đăng ký'),
    'TC_NAV_005':       ('NAV_05',   'Điều hướng từ đăng ký sang đăng nhập'),
    'TC_NAV_006':       ('NAV_06',   'Xác minh tiêu đề cửa sổ trình duyệt'),
    'TC_NAV_007':       ('NAV_07',   'Kiểm tra ứng dụng khi thay đổi kích thước cửa sổ'),
    'TC_NAV_008':       ('NAV_08',   'URL gốc chuyển hướng theo trạng thái xác thực'),
}

# ── Đọc report mới nhất ─────────────────────────────────────────────────────
reports = sorted(glob.glob(r'reports\selenium_test_report_*.xlsx'), key=os.path.getctime)
latest_report = reports[-1]
print(f"Đọc report: {latest_report}")
rpt_wb = openpyxl.load_workbook(latest_report)

# ── Tạo file report mới với Test ID đúng ────────────────────────────────────
out_name = f"reports\\selenium_test_report_updated.xlsx"
shutil.copy(latest_report, out_name)
out_wb = openpyxl.load_workbook(out_name)

ws = out_wb['Detailed Results']
# Cập nhật cột Test ID (cột B = index 2)
rows_updated = 0
for row in ws.iter_rows(min_row=2):
    old_id = row[1].value  # cột Test ID
    if old_id and old_id in ID_MAP:
        new_id, desc = ID_MAP[old_id]
        row[1].value = new_id   # Gán Test ID mới từ system file
        rows_updated += 1
        print(f"  {old_id} -> {new_id} | {desc}")

out_wb.save(out_name)
print(f"\nĐã cập nhật {rows_updated} dòng -> {out_name}")

# ── Cập nhật 02_system - Copy.xlsx ─────────────────────────────────────────
SYSTEM_FILE = r'..\02_system - Copy.xlsx'
sys_wb = openpyxl.load_workbook(SYSTEM_FILE)

# Kết quả từ report
results = {}
rpt_ws = rpt_wb['Detailed Results']
for row in rpt_ws.iter_rows(min_row=2, values_only=True):
    if row[1] and row[5]:
        results[row[1]] = row[5]  # old_id -> status

# Ánh xạ ngược: system_id -> status
sys_results = {}
for old_id, status in results.items():
    if old_id in ID_MAP:
        new_id = ID_MAP[old_id][0]
        sys_results[new_id] = 'Pass' if status == 'PASS' else 'Fail'

print("\n=== Cập nhật 02_system - Copy.xlsx ===")

# Cập nhật Feature_IM
for sheet_name in ['Feature_IM', 'Feature_RL']:
    ws_sys = sys_wb[sheet_name]
    for row in ws_sys.iter_rows(min_row=10):
        test_id_cell = row[1]  # Cột B = Test ID
        result_cell = row[26]  # Cột AC = Testing Result (index 26)
        if test_id_cell.value and str(test_id_cell.value).strip() in sys_results:
            tid = str(test_id_cell.value).strip()
            result_cell.value = sys_results[tid]
            print(f"  [{sheet_name}] {tid} -> {sys_results[tid]}")

sys_wb.save(SYSTEM_FILE)
print(f"\nĐã lưu: {SYSTEM_FILE}")
print("\nHoàn thành!")
