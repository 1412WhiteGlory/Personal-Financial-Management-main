"""
Excel Report Generator for Selenium Test Results.
Generates a professional Excel report with multiple sheets:
  - Summary sheet with pass/fail statistics
  - Detailed results sheet with all test outcomes
  - Per-module breakdown
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

import config


def generate_report(test_results, output_path=None):
    """
    Generate an Excel report from collected test results.
    
    Args:
        test_results: list of dicts with keys:
            test_id, test_name, module, description, status,
            actual_result, error_message, screenshot, duration_sec, timestamp
        output_path: path for the output Excel file
    """
    output_path = output_path or config.REPORT_PATH
    wb = Workbook()

    # ────────────────────────────────────────────────────────
    # Styles
    # ────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    skip_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    subtitle_font = Font(name="Calibri", bold=True, size=11, color="404040")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def apply_border(ws, min_row, max_row, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).border = thin_border

    # ────────────────────────────────────────────────────────
    # Sheet 1: Summary
    # ────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    total = len(test_results)
    passed = sum(1 for r in test_results if r["status"] == "PASS")
    failed = sum(1 for r in test_results if r["status"] == "FAIL")
    errors = sum(1 for r in test_results if r["status"] == "ERROR")
    skipped = sum(1 for r in test_results if r["status"] == "SKIP")
    pass_rate = (passed / total * 100) if total > 0 else 0
    total_duration = sum(r["duration_sec"] for r in test_results)

    ws_summary.cell(row=1, column=1, value="SELENIUM WEBDRIVER TEST REPORT").font = title_font
    ws_summary.cell(row=2, column=1, value="Personal Financial Management Application").font = subtitle_font
    ws_summary.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = subtitle_font

    summary_data = [
        ("Metric", "Value"),
        ("Total Test Cases", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Errors", errors),
        ("Skipped", skipped),
        ("Pass Rate (%)", f"{pass_rate:.1f}%"),
        ("Total Duration (sec)", f"{total_duration:.1f}"),
    ]

    for i, (label, value) in enumerate(summary_data, start=5):
        ws_summary.cell(row=i, column=1, value=label)
        ws_summary.cell(row=i, column=2, value=value)
        if i == 5:
            style_header_row(ws_summary, i, 2)
        else:
            ws_summary.cell(row=i, column=1).font = Font(bold=True)
            ws_summary.cell(row=i, column=1).border = thin_border
            ws_summary.cell(row=i, column=2).border = thin_border
            # Color code
            if label == "Passed":
                ws_summary.cell(row=i, column=2).fill = pass_fill
            elif label == "Failed":
                ws_summary.cell(row=i, column=2).fill = fail_fill
            elif label == "Errors":
                ws_summary.cell(row=i, column=2).fill = error_fill

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20

    # Pie Chart
    if total > 0:
        chart_data_start = 15
        ws_summary.cell(row=chart_data_start, column=1, value="Status")
        ws_summary.cell(row=chart_data_start, column=2, value="Count")
        for i, (label, count) in enumerate([("PASS", passed), ("FAIL", failed),
                                             ("ERROR", errors), ("SKIP", skipped)], start=1):
            ws_summary.cell(row=chart_data_start + i, column=1, value=label)
            ws_summary.cell(row=chart_data_start + i, column=2, value=count)

        pie = PieChart()
        pie.title = "Test Results Distribution"
        pie.style = 10
        labels = Reference(ws_summary, min_col=1, min_row=chart_data_start + 1,
                           max_row=chart_data_start + 4)
        data = Reference(ws_summary, min_col=2, min_row=chart_data_start,
                         max_row=chart_data_start + 4)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 15
        pie.height = 10
        ws_summary.add_chart(pie, "D5")

    # ────────────────────────────────────────────────────────
    # Sheet 2: Detailed Results
    # ────────────────────────────────────────────────────────
    ws_detail = wb.create_sheet("Detailed Results")

    headers = [
        "No.", "Test ID", "Test Name", "Module", "Description",
        "Status", "Actual Result", "Error Message", "Screenshot",
        "Duration (s)", "Timestamp"
    ]
    for col, h in enumerate(headers, start=1):
        ws_detail.cell(row=1, column=col, value=h)
    style_header_row(ws_detail, 1, len(headers))

    for idx, result in enumerate(test_results, start=1):
        row = idx + 1
        ws_detail.cell(row=row, column=1, value=idx)
        ws_detail.cell(row=row, column=2, value=result.get("test_id", ""))
        ws_detail.cell(row=row, column=3, value=result.get("test_name", ""))
        ws_detail.cell(row=row, column=4, value=result.get("module", ""))
        ws_detail.cell(row=row, column=5, value=result.get("description", ""))

        status_cell = ws_detail.cell(row=row, column=6, value=result.get("status", ""))
        if result["status"] == "PASS":
            status_cell.fill = pass_fill
        elif result["status"] == "FAIL":
            status_cell.fill = fail_fill
        elif result["status"] == "ERROR":
            status_cell.fill = error_fill
        elif result["status"] == "SKIP":
            status_cell.fill = skip_fill

        ws_detail.cell(row=row, column=7, value=result.get("actual_result", ""))
        ws_detail.cell(row=row, column=8, value=result.get("error_message", ""))
        ws_detail.cell(row=row, column=9, value=result.get("screenshot", ""))
        ws_detail.cell(row=row, column=10, value=result.get("duration_sec", 0))
        ws_detail.cell(row=row, column=11, value=result.get("timestamp", ""))

    apply_border(ws_detail, 1, len(test_results) + 1, len(headers))

    # Adjust column widths
    col_widths = [5, 15, 25, 20, 40, 10, 50, 50, 40, 12, 22]
    for i, w in enumerate(col_widths, start=1):
        ws_detail.column_dimensions[get_column_letter(i)].width = w

    # ────────────────────────────────────────────────────────
    # Sheet 3: Module Breakdown
    # ────────────────────────────────────────────────────────
    ws_module = wb.create_sheet("Module Breakdown")

    module_stats = {}
    for r in test_results:
        mod = r.get("module", "Unknown")
        if mod not in module_stats:
            module_stats[mod] = {"total": 0, "pass": 0, "fail": 0, "error": 0}
        module_stats[mod]["total"] += 1
        if r["status"] == "PASS":
            module_stats[mod]["pass"] += 1
        elif r["status"] == "FAIL":
            module_stats[mod]["fail"] += 1
        elif r["status"] == "ERROR":
            module_stats[mod]["error"] += 1

    mod_headers = ["Module", "Total", "Pass", "Fail", "Error", "Pass Rate (%)"]
    for col, h in enumerate(mod_headers, start=1):
        ws_module.cell(row=1, column=col, value=h)
    style_header_row(ws_module, 1, len(mod_headers))

    for idx, (mod, stats) in enumerate(sorted(module_stats.items()), start=2):
        pr = (stats["pass"] / stats["total"] * 100) if stats["total"] > 0 else 0
        ws_module.cell(row=idx, column=1, value=mod)
        ws_module.cell(row=idx, column=2, value=stats["total"])
        ws_module.cell(row=idx, column=3, value=stats["pass"]).fill = pass_fill
        ws_module.cell(row=idx, column=4, value=stats["fail"]).fill = fail_fill if stats["fail"] > 0 else PatternFill()
        ws_module.cell(row=idx, column=5, value=stats["error"]).fill = error_fill if stats["error"] > 0 else PatternFill()
        ws_module.cell(row=idx, column=6, value=f"{pr:.1f}%")

    apply_border(ws_module, 1, len(module_stats) + 1, len(mod_headers))

    mod_col_widths = [25, 10, 10, 10, 10, 15]
    for i, w in enumerate(mod_col_widths, start=1):
        ws_module.column_dimensions[get_column_letter(i)].width = w

    # ────────────────────────────────────────────────────────
    # Sheet 4: Test Data Info
    # ────────────────────────────────────────────────────────
    ws_data = wb.create_sheet("Test Data")
    ws_data.cell(row=1, column=1, value="Test Data Files Used").font = title_font
    data_files = [
        ("test_users.csv", "User registration and login test data"),
        ("test_income.csv", "Income CRUD test data"),
        ("test_expense.csv", "Expense CRUD test data"),
    ]
    data_headers = ["File", "Description"]
    for col, h in enumerate(data_headers, start=1):
        ws_data.cell(row=3, column=col, value=h)
    style_header_row(ws_data, 3, 2)

    for idx, (f, desc) in enumerate(data_files, start=4):
        ws_data.cell(row=idx, column=1, value=f)
        ws_data.cell(row=idx, column=2, value=desc)

    ws_data.column_dimensions["A"].width = 25
    ws_data.column_dimensions["B"].width = 50

    # ────────────────────────────────────────────────────────
    # Save
    # ────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n{'=' * 60}")
    print(f"📊 TEST REPORT GENERATED: {output_path}")
    print(f"   Total: {total} | Pass: {passed} | Fail: {failed} | Error: {errors}")
    print(f"   Pass Rate: {pass_rate:.1f}%")
    print(f"   Duration: {total_duration:.1f}s")
    print(f"{'=' * 60}\n")
    return output_path


if __name__ == "__main__":
    # Quick test with sample data
    sample = [
        {
            "test_id": "TC_001", "test_name": "Sample Test", "module": "Sample",
            "description": "A sample test", "status": "PASS", "actual_result": "OK",
            "error_message": "", "screenshot": "", "duration_sec": 1.5,
            "timestamp": "2025-04-22 10:00:00"
        }
    ]
    generate_report(sample)
