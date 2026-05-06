import os
import glob
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def main():
    # 1. Find the latest test report
    report_files = glob.glob('reports/selenium_test_report_*.xlsx')
    if not report_files:
        print("No test reports found.")
        return
    latest_report = max(report_files, key=os.path.getctime)
    print(f"Reading {latest_report}...")
    
    wb_results = openpyxl.load_workbook(latest_report)
    ws_results = wb_results['Detailed Results']
    
    test_results = {}
    for row in ws_results.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        test_id = row[1]
        test_results[test_id] = {
            'test_name': row[2],
            'module': row[3],
            'description': row[4],
            'status': row[5],
            'actual_result': row[6],
            'error_message': row[7],
            'screenshot': row[8],
            'duration': row[9],
            'timestamp': row[10]
        }
        
    # 2. Generate new report like Sample_tool_test.xlsx
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Test Report"
    headers = ['Test Case ID', 'Test type', 'Objective', 'Steps', 'Input', 'Expected output', 'Tester', 'Test date', 'Status', 'Notes']
    ws_new.append(headers)
    
    for test_id, data in test_results.items():
        ws_new.append([
            test_id,
            data['module'],
            data['description'],
            '', # Steps
            '', # Input
            '', # Expected output
            'AutoTester',
            data['timestamp'],
            data['status'],
            data['screenshot'] if data['screenshot'] else data['error_message']
        ])
    
    new_report_name = f"Final_Test_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb_new.save(new_report_name)
    print(f"Generated new report: {new_report_name}")
    
    # 3. Update 02_system - Copy.xlsx
    sys_file = '../02_system - Copy.xlsx'
    wb_sys = openpyxl.load_workbook(sys_file)
    
    for sheet_name in wb_sys.sheetnames:
        ws = wb_sys[sheet_name]
        
        # Find column indices for Evaluation and Testing Result
        eval_col = None
        result_col = None
        for row in ws.iter_rows(min_row=1, max_row=15):
            for cell in row:
                val = str(cell.value).strip().lower() if cell.value else ""
                if val == "evaluation" or val == "đánh giá":
                    eval_col = cell.column
                elif val == "testing result" or val == "kết quả":
                    result_col = cell.column
        
        if not eval_col:
            eval_col = 29 # fallback
        if not result_col:
            result_col = 23 # fallback
            
        for row in ws.iter_rows(min_row=1):
            cell_id = row[1] # Column B
            if cell_id.value and str(cell_id.value).strip() in test_results:
                t_id = str(cell_id.value).strip()
                status = test_results[t_id]['status']
                ss = test_results[t_id]['screenshot']
                
                eval_cell = ws.cell(row=cell_id.row, column=eval_col)
                eval_cell.value = "Pass" if status == "PASS" else "Fail"
                
                # Highlight
                if status == "PASS":
                    eval_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    eval_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                result_cell = ws.cell(row=cell_id.row, column=result_col)
                if ss:
                    result_cell.value = ss
                    
    wb_sys.save(sys_file)
    print(f"Updated {sys_file}")

if __name__ == '__main__':
    main()
