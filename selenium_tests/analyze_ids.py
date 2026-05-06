"""
Script tổng hợp:
1. Đọc 02_system - Copy.xlsx để lấy Test ID chính xác
2. Cập nhật report Excel mới nhất với đúng Test ID
3. Tổng hợp mapping để update test files
"""
import openpyxl, glob, os, sys
sys.stdout = __import__('io').TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SYSTEM_FILE = r'f:\HOC\DATN_ANH_QUAN\Personal-Financial-Management-main (1)\Personal-Financial-Management-main-clone\02_system - Copy.xlsx'

wb = openpyxl.load_workbook(SYSTEM_FILE)

print("=== Feature_IM Test IDs ===")
ws = wb['Feature_IM']
for row in ws.iter_rows(min_row=16, values_only=True):
    if row[1] and str(row[1]).startswith('IM'):
        print(f"  {row[1]} | {str(row[3])[:70]}")

print("\n=== Feature_RL Test IDs ===")
ws = wb['Feature_RL']
for row in ws.iter_rows(min_row=16, values_only=True):
    if row[1] and (str(row[1]).startswith('REG') or str(row[1]).startswith('LOG') or str(row[1]).startswith('INT')):
        print(f"  {row[1]} | {str(row[3])[:70]}")

# Also read current test report
print("\n=== Latest Test Report ===")
reports = sorted(glob.glob(r'reports\selenium_test_report_*.xlsx'), key=os.path.getctime)
if reports:
    rpt = openpyxl.load_workbook(reports[-1])
    ws_det = rpt['Detailed Results']
    print("Headers:", [c.value for c in list(ws_det.iter_rows(min_row=1, max_row=1))[0]])
    for row in ws_det.iter_rows(min_row=2, values_only=True):
        if row[1]:
            print(f"  {row[1]} | {row[2]} | {row[5]}")
