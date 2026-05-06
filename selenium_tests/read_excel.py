import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'f:\HOC\DATN_ANH_QUAN\Personal-Financial-Management-main (1)\Personal-Financial-Management-main-clone\02_system - Copy.xlsx')
print("=== Sheets ===")
print(wb.sheetnames)

# Focus on Feature_IM sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (max_row={ws.max_row}) ===")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(cell is not None for cell in row):
            print(f"  R{i+1}: {[str(c)[:40] if c else None for c in row[:10]]}")
        if i > 60:
            print("  ...(truncated)")
            break
