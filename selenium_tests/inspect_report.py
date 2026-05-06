import openpyxl, sys, io, glob, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'reports\selenium_test_report.xlsx')
print("Sheets:", wb.sheetnames)

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== Sheet: {sn} | rows={ws.max_row} cols={ws.max_column} ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        print(f"  R{i+1}: {list(row)}")
