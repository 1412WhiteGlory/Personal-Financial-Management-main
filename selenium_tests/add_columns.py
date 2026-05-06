# -*- coding: utf-8 -*-
"""Thêm cột Steps, Input, Expected Output vào selenium_test_report.xlsx"""
import openpyxl, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Dữ liệu cho từng Test ID ──────────────────────────────────────────────────
DATA = {
# ═══ ĐĂNG KÝ ═══
"REG_01": {
    "steps": "1. Mở trình duyệt, truy cập /signUp\n2. Nhập họ tên hợp lệ\n3. Nhập email hợp lệ (duy nhất)\n4. Nhập mật khẩu đáp ứng yêu cầu\n5. Nhấp nút Đăng ký\n6. Kiểm tra chuyển hướng sang /dashboard\n7. Gọi API /auth/login xác minh tài khoản trong DB",
    "input": "fullName: 'Selenium Test User'\nemail: 'test_reg001_{timestamp}@test.com'\npassword: 'SeleniumTest@123'",
    "expected": "- Chuyển hướng đến /dashboard\n- API login trả về HTTP 200\n- Token hợp lệ được cấp",
},
"REG_02": {
    "steps": "1. Mở trình duyệt, truy cập /signUp\n2. Để trống trường họ tên\n3. Nhập email và mật khẩu hợp lệ\n4. Nhấp nút Đăng ký\n5. Kiểm tra trang không chuyển hướng\n6. Kiểm tra thông báo lỗi xuất hiện",
    "input": "fullName: '' (trống)\nemail: 'test_missing_name@test.com'\npassword: 'SeleniumTest@123'",
    "expected": "- Ở lại trang /signUp\n- Hiển thị thông báo lỗi trường bắt buộc\n- Không tạo tài khoản mới",
},
"REG_03": {
    "steps": "1. Truy cập /signUp\n2. Nhập họ tên hợp lệ\n3. Nhập email sai định dạng (thiếu @)\n4. Nhập mật khẩu hợp lệ\n5. Nhấp Đăng ký\n6. Kiểm tra lỗi email",
    "input": "fullName: 'Test User'\nemail: 'invalidemail' (không có @)\npassword: 'SeleniumTest@123'",
    "expected": "- Ở lại trang /signUp\n- Thông báo 'email không hợp lệ'\n- Không tạo tài khoản",
},
"REG_05": {
    "steps": "1. Truy cập /signUp\n2. Nhập họ tên, email hợp lệ\n3. Nhập mật khẩu không đủ yêu cầu\n4. Nhấp Đăng ký\n5. Kiểm tra thông báo lỗi mật khẩu",
    "input": "fullName: 'Test User'\nemail: 'test@test.com'\npassword: 'abc' (quá ngắn) / 'abcdefgh1!' (không có chữ hoa) / 'Abcdefgh1' (không có ký tự đặc biệt)",
    "expected": "- Ở lại trang /signUp\n- Thông báo yêu cầu mật khẩu (≥8 ký tự, có chữ hoa, có ký tự đặc biệt)\n- Không tạo tài khoản",
},
"INT_01": {
    "steps": "1. Truy cập /signUp\n2. Kiểm tra tiêu đề trang\n3. Đếm số trường input (≥3)\n4. Kiểm tra nút submit tồn tại\n5. Kiểm tra liên kết 'Đã có tài khoản? Đăng nhập'",
    "input": "Không nhập dữ liệu, chỉ quan sát giao diện",
    "expected": "- Tiêu đề 'Tạo tài khoản' / 'Sign Up' hiển thị\n- Có ≥3 trường input (họ tên, email, mật khẩu)\n- Nút submit hiển thị\n- Liên kết đến /login tồn tại",
},
# ═══ ĐĂNG NHẬP ═══
"LOG_01": {
    "steps": "1. Xóa localStorage, truy cập /login\n2. Nhập email tài khoản test\n3. Nhập mật khẩu đúng\n4. Nhấp nút Đăng nhập\n5. Kiểm tra chuyển hướng /dashboard\n6. Lấy token từ localStorage\n7. Gọi API /auth/getUser xác minh token",
    "input": "email: 'selenium_test_user@test.com'\npassword: 'SeleniumTest@123'",
    "expected": "- Chuyển hướng đến /dashboard\n- localStorage có 'token'\n- API /auth/getUser trả về HTTP 200",
},
"LOG_02": {
    "steps": "1. Truy cập /login\n2. Nhập email/mật khẩu sai hoặc trống\n3. Nhấp Đăng nhập\n4. Kiểm tra trang không chuyển hướng\n5. Kiểm tra thông báo lỗi",
    "input": "Trường hợp 1: password='WrongPassword@999'\nTrường hợp 2: email='' (trống)\nTrường hợp 3: password='' (trống)",
    "expected": "- Ở lại trang /login\n- Thông báo lỗi xác thực hiển thị\n- Không có token trong localStorage",
},
"LOG_03": {
    "steps": "1. Truy cập /login\n2. Nhập email chưa đăng ký trong hệ thống\n3. Nhập mật khẩu bất kỳ\n4. Nhấp Đăng nhập\n5. Kiểm tra lỗi",
    "input": "email: 'nonexistent_user_xyz@nowhere.com'\npassword: 'Test@1234'",
    "expected": "- Ở lại trang /login\n- Thông báo 'thông tin không hợp lệ'\n- Không có token",
},
"LOG_04": {
    "steps": "1. Truy cập /login\n2. Tìm liên kết 'Quên mật khẩu'\n3. Nhấp liên kết\n4. Kiểm tra URL chứa 'forgot'",
    "input": "Nhấp liên kết 'Quên mật khẩu?' trên trang đăng nhập",
    "expected": "- Chuyển hướng đến /forgot-password\n- Trang đặt lại mật khẩu hiển thị",
},
"INT_02": {
    "steps": "1. Truy cập /login\n2. Kiểm tra tiêu đề trang\n3. Đếm trường input (≥2)\n4. Kiểm tra nút submit\n5. Kiểm tra liên kết đăng ký và quên mật khẩu",
    "input": "Không nhập dữ liệu, chỉ quan sát giao diện",
    "expected": "- Tiêu đề 'Đăng nhập' / 'Chào mừng' hiển thị\n- Có ≥2 trường (email, mật khẩu)\n- Nút đăng nhập hiển thị\n- Liên kết đến /signUp tồn tại",
},
# ═══ THU NHẬP ═══
"IM_IC_03": {
    "steps": "1. Đăng nhập, truy cập /income\n2. Ghi số lượng bản ghi hiện tại qua API\n3. Nhấp nút 'Thêm thu nhập'\n4. Điền nguồn thu nhập\n5. Điền số tiền\n6. Điền ngày thu nhập (dùng JS)\n7. Nhấp 'Add Income'\n8. Gọi API kiểm tra số lượng tăng lên 1",
    "input": "source: 'Salary' / 'Delete Test'\namount: 5000000 / 100000\ndate: '2025-04-01' / '2025-04-20'",
    "expected": "- Modal đóng sau khi submit\n- API GET /income trả về count+1\n- Bản ghi mới xuất hiện trong danh sách",
},
"IM_IC_05": {
    "steps": "1. Đăng nhập, truy cập /income\n2. Ghi count trước qua API\n3. Nhấp 'Thêm thu nhập'\n4. Để trống trường Nguồn\n5. Điền số tiền và ngày\n6. Nhấp 'Add Income'\n7. Kiểm tra count không đổi hoặc có lỗi",
    "input": "source: '' (trống)\namount: 100000\ndate: '2025-04-01'",
    "expected": "- Count DB không tăng\n- Thông báo lỗi 'trường bắt buộc'\n- Modal không đóng",
},
"IM_IC_06": {
    "steps": "1. Đăng nhập, truy cập /income\n2. Ghi count trước qua API\n3. Nhấp 'Thêm thu nhập'\n4. Điền nguồn hợp lệ\n5. Nhập số tiền âm hoặc 0\n6. Nhấp 'Add Income'\n7. Kiểm tra count không đổi",
    "input": "source: 'Test Income'\namount: -1000 (âm)\ndate: '2025-04-01'",
    "expected": "- Count DB không tăng\n- Thông báo 'số tiền phải lớn hơn 0'\n- Modal không đóng",
},
"IM_IC_07": {
    "steps": "1. Đăng nhập, truy cập /income\n2. Ghi count trước qua API\n3. Thêm thu nhập freelance qua UI\n4. Nhấp 'Add Income'\n5. Gọi API kiểm tra count tăng\n6. Kiểm tra danh sách UI cập nhật",
    "input": "source: 'Freelance Project'\namount: 2000000\ndate: '2025-04-10'",
    "expected": "- API count tăng +1\n- Bản ghi 'Freelance Project' xuất hiện trong danh sách\n- Không có lỗi",
},
"IM_01": {
    "steps": "1. Đăng nhập, truy cập /income\n2. Kiểm tra nút 'Thêm thu nhập'\n3. Kiểm tra bộ lọc ngày\n4. Kiểm tra danh sách thu nhập\n5. Kiểm tra thanh điều hướng sidebar",
    "input": "Không nhập dữ liệu, chỉ quan sát giao diện trang /income",
    "expected": "- Nút 'Thêm thu nhập' hiển thị\n- Bộ lọc date picker hiển thị\n- Danh sách bản ghi hiển thị\n- Sidebar có các liên kết điều hướng",
},
# ═══ CHI TIÊU ═══
"EXP_IC_03": {
    "steps": "1. Đăng nhập, truy cập /expense\n2. Ghi count trước qua API\n3. Nhấp 'Thêm chi tiêu'\n4. Điền danh mục\n5. Điền số tiền\n6. Điền ngày (dùng JS)\n7. Nhấp submit\n8. Kiểm tra count tăng qua API",
    "input": "category: 'Food' / 'Delete Test'\namount: 150000 / 50000\ndate: '2025-04-05' / '2025-04-20'",
    "expected": "- Modal đóng sau submit\n- API GET /expense count+1\n- Bản ghi mới xuất hiện trong danh sách",
},
"EXP_IC_05": {
    "steps": "1. Đăng nhập, truy cập /expense\n2. Ghi count trước\n3. Nhấp 'Thêm chi tiêu'\n4. Để trống trường Danh mục\n5. Điền số tiền và ngày\n6. Nhấp submit\n7. Kiểm tra count không đổi",
    "input": "category: '' (trống)\namount: 100000\ndate: '2025-04-05'",
    "expected": "- Count DB không tăng\n- Thông báo lỗi 'danh mục bắt buộc'\n- Modal không đóng",
},
"EXP_IC_06": {
    "steps": "1. Đăng nhập, truy cập /expense\n2. Ghi count trước\n3. Nhấp 'Thêm chi tiêu'\n4. Điền danh mục hợp lệ\n5. Nhập số tiền âm\n6. Nhấp submit\n7. Kiểm tra count không đổi",
    "input": "category: 'Test Expense'\namount: -500 (âm)\ndate: '2025-04-05'",
    "expected": "- Count DB không tăng\n- Thông báo 'số tiền phải lớn hơn 0'\n- Modal không đóng",
},
"EXP_IC_07": {
    "steps": "1. Đăng nhập, truy cập /expense\n2. Ghi count trước\n3. Thêm chi tiêu giao thông qua UI\n4. Nhấp submit\n5. Kiểm tra count tăng\n6. Kiểm tra danh sách cập nhật",
    "input": "category: 'Transport'\namount: 80000\ndate: '2025-04-08'",
    "expected": "- API count tăng +1\n- Bản ghi 'Transport' xuất hiện trong danh sách\n- Không có lỗi",
},
"EXP_01": {
    "steps": "1. Đăng nhập, truy cập /expense\n2. Kiểm tra nút 'Thêm chi tiêu'\n3. Kiểm tra bộ lọc ngày\n4. Kiểm tra danh sách chi tiêu\n5. Kiểm tra sidebar",
    "input": "Không nhập dữ liệu, chỉ quan sát giao diện trang /expense",
    "expected": "- Nút 'Thêm chi tiêu' hiển thị\n- Bộ lọc date picker hiển thị\n- Danh sách bản ghi hiển thị\n- Sidebar điều hướng hoạt động",
},
# ═══ DASHBOARD ═══
"DASH_01": {
    "steps": "1. Đăng nhập qua UI\n2. Truy cập /dashboard\n3. Kiểm tra URL chứa 'dashboard'\n4. Kiểm tra nội dung tài chính hiển thị",
    "input": "email: 'selenium_test_user@test.com'\npassword: 'SeleniumTest@123'",
    "expected": "- URL = http://localhost:5173/dashboard\n- Trang có nội dung tài chính (Thu nhập, Chi tiêu, Số dư)",
},
"DASH_02": {
    "steps": "1. Đăng nhập, truy cập /dashboard\n2. Kiểm tra thẻ 'Tổng số dư'\n3. Kiểm tra thẻ 'Tổng thu nhập'\n4. Kiểm tra thẻ 'Tổng chi tiêu'",
    "input": "Dữ liệu đã seed: Thu nhập 10.000.000đ, Chi tiêu 3.000.000đ",
    "expected": "- 3 thẻ thông tin hiển thị đầy đủ\n- Số liệu khớp với dữ liệu đã thêm vào DB",
},
"DASH_03": {
    "steps": "1. Gọi API GET /dashboard lấy dữ liệu\n2. Đăng nhập, truy cập /dashboard\n3. So sánh totalIncome và totalExpenses trên UI với API",
    "input": "API call: GET /api/v1/dashboard với Bearer token",
    "expected": "- API trả về HTTP 200\n- Dữ liệu UI nhất quán với dữ liệu API\n- totalIncome và totalExpenses hiển thị đúng",
},
"DASH_04": {
    "steps": "1. Đăng nhập, truy cập /dashboard\n2. Tìm liên kết/nút 'Thu nhập' trong sidebar\n3. Nhấp liên kết\n4. Kiểm tra URL chứa 'income'",
    "input": "Nhấp nút/liên kết 'Thu nhập' trên sidebar hoặc trang dashboard",
    "expected": "- URL chuyển sang /income\n- Trang quản lý thu nhập hiển thị",
},
"DASH_05": {
    "steps": "1. Đăng nhập, truy cập /dashboard\n2. Tìm liên kết/nút 'Chi tiêu'\n3. Nhấp liên kết\n4. Kiểm tra URL chứa 'expense'",
    "input": "Nhấp nút/liên kết 'Chi tiêu' trên sidebar hoặc trang dashboard",
    "expected": "- URL chuyển sang /expense\n- Trang quản lý chi tiêu hiển thị",
},
"DASH_06": {
    "steps": "1. Đăng nhập, truy cập /dashboard\n2. Chờ trang render hoàn chỉnh (3 giây)\n3. Tìm phần tử SVG (Recharts) hoặc Canvas\n4. Kiểm tra số lượng > 0",
    "input": "Dữ liệu đã seed vào DB để biểu đồ có dữ liệu hiển thị",
    "expected": "- Có ≥1 phần tử SVG hoặc Canvas\n- Biểu đồ thu nhập/chi tiêu render thành công",
},
"DASH_07": {
    "steps": "1. Đăng nhập, truy cập /dashboard\n2. Tìm section 'Giao dịch gần đây'\n3. Kiểm tra text 'Giao dịch' / 'Transaction' / 'gần đây' trong trang",
    "input": "Dữ liệu đã seed: có bản ghi thu nhập và chi tiêu gần đây",
    "expected": "- Section 'Giao dịch gần đây' hiển thị\n- Danh sách giao dịch có ít nhất 1 mục",
},
"DASH_UI_01": {
    "steps": "1. Đăng nhập, truy cập /dashboard\n2. Kiểm tra phần tử grid layout\n3. Lấy tiêu đề cửa sổ trình duyệt\n4. Kiểm tra số grid element > 0",
    "input": "Không nhập dữ liệu, kiểm tra cấu trúc DOM",
    "expected": "- Có ≥1 phần tử class chứa 'grid'\n- Tiêu đề cửa sổ không trống (title='PFM')",
},
# ═══ ĐIỀU HƯỚNG ═══
"NAV_01": {
    "steps": "1. Xóa localStorage (đảm bảo chưa đăng nhập)\n2. Truy cập trực tiếp /dashboard\n3. Chờ 2 giây\n4. Kiểm tra URL hiện tại chứa 'login'",
    "input": "URL: http://localhost:5173/dashboard\nTrạng thái: chưa đăng nhập",
    "expected": "- Tự động chuyển hướng về /login\n- Không cho truy cập dashboard khi chưa xác thực",
},
"NAV_02": {
    "steps": "1. Xóa localStorage\n2. Truy cập trực tiếp /income\n3. Chờ 2 giây\n4. Kiểm tra URL chứa 'login'",
    "input": "URL: http://localhost:5173/income\nTrạng thái: chưa đăng nhập",
    "expected": "- Chuyển hướng về /login\n- Route protection hoạt động đúng",
},
"NAV_03": {
    "steps": "1. Đăng nhập, truy cập /dashboard\n2. Tìm nút sidebar (Trang chủ/Dashboard)\n3. Nhấp nút\n4. Kiểm tra URL cập nhật đúng",
    "input": "Nhấp các nút điều hướng trong sidebar: Dashboard, Thu nhập, Chi tiêu",
    "expected": "- URL thay đổi tương ứng (/dashboard, /income, /expense)\n- Trang đích tải thành công",
},
"NAV_04": {
    "steps": "1. Truy cập /login\n2. Tìm liên kết 'Đăng ký'\n3. Nhấp liên kết\n4. Kiểm tra URL chứa 'signup'",
    "input": "Nhấp liên kết 'Chưa có tài khoản? Đăng ký' trên trang /login",
    "expected": "- Chuyển hướng đến /signUp\n- Trang đăng ký hiển thị",
},
"NAV_05": {
    "steps": "1. Truy cập /signUp\n2. Tìm liên kết 'Đăng nhập'\n3. Nhấp liên kết\n4. Kiểm tra URL chứa 'login'",
    "input": "Nhấp liên kết 'Đã có tài khoản? Đăng nhập' trên trang /signUp",
    "expected": "- Chuyển hướng đến /login\n- Trang đăng nhập hiển thị",
},
"NAV_06": {
    "steps": "1. Truy cập /dashboard (đã đăng nhập)\n2. Lấy driver.title\n3. Kiểm tra title không rỗng",
    "input": "Không nhập dữ liệu",
    "expected": "- Tiêu đề cửa sổ trình duyệt = 'PFM' (không rỗng)",
},
"NAV_07": {
    "steps": "1. Đăng nhập, truy cập /dashboard\n2. Thay đổi kích thước cửa sổ: 1920x1080\n3. Thay đổi: 1366x768\n4. Thay đổi: 768x1024 (tablet)\n5. Thay đổi: 375x812 (mobile)\n6. Kiểm tra trang không crash ở mỗi kích thước",
    "input": "Viewport sizes: [(1920,1080), (1366,768), (768,1024), (375,812)]",
    "expected": "- Trang tải thành công ở mọi kích thước\n- Không có lỗi JavaScript\n- URL vẫn là /dashboard",
},
"NAV_08": {
    "steps": "1. Xóa localStorage, truy cập /\n2. Kiểm tra chuyển hướng về /login\n3. Đăng nhập thành công\n4. Truy cập /\n5. Kiểm tra chuyển hướng về /dashboard",
    "input": "URL: http://localhost:5173/\nTrường hợp 1: chưa đăng nhập\nTrường hợp 2: đã đăng nhập",
    "expected": "- Chưa đăng nhập: / → /login\n- Đã đăng nhập: / → /dashboard",
},
}

# ── Xử lý file Excel ──────────────────────────────────────────────────────────
SRC = r'reports\selenium_test_report_updated.xlsx'
OUT = r'reports\selenium_test_report.xlsx'

wb = openpyxl.load_workbook(SRC)
ws = wb['Detailed Results']

# Tìm hàng header (hàng 1)
header_row = 1
max_col = ws.max_column  # cột cuối hiện tại

# Thêm 3 tiêu đề cột mới
new_cols = ['Steps', 'Input', 'Expected Output']
col_steps = max_col + 1
col_input = max_col + 2
col_expected = max_col + 3

# Style cho header mới
hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

cell_align = Alignment(vertical="top", wrap_text=True)

for i, col_name in enumerate(new_cols, start=col_steps):
    cell = ws.cell(row=header_row, column=i, value=col_name)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

# Điền dữ liệu vào từng hàng dựa theo Test ID (cột B = cột 2)
test_id_col = 2  # cột B chứa Test ID
filled = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    tid = row[test_id_col - 1].value
    if tid and str(tid).strip() in DATA:
        d = DATA[str(tid).strip()]
        steps_cell = ws.cell(row=row[0].row, column=col_steps, value=d["steps"])
        input_cell = ws.cell(row=row[0].row, column=col_input, value=d["input"])
        exp_cell   = ws.cell(row=row[0].row, column=col_expected, value=d["expected"])
        for c in [steps_cell, input_cell, exp_cell]:
            c.alignment = cell_align
            c.border = border
        filled += 1
        print(f"  OK: {tid}")

# Điều chỉnh độ rộng cột mới
ws.column_dimensions[get_column_letter(col_steps)].width = 60
ws.column_dimensions[get_column_letter(col_input)].width = 40
ws.column_dimensions[get_column_letter(col_expected)].width = 45

# Điều chỉnh chiều cao hàng cho dễ đọc
for row_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 90

wb.save(OUT)
print(f"\nĐã thêm 3 cột vào {filled} dòng -> {OUT}")
