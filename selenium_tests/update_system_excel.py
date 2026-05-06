"""
Cập nhật 02_system - Copy.xlsx với kết quả test - xử lý merged cells
"""
import openpyxl, glob, os
from openpyxl.utils import get_column_letter

SYSTEM_FILE = r'f:\HOC\DATN_ANH_QUAN\Personal-Financial-Management-main (1)\Personal-Financial-Management-main-clone\02_system - Copy.xlsx'
OUT_FILE = r'f:\HOC\DATN_ANH_QUAN\Personal-Financial-Management-main (1)\Personal-Financial-Management-main-clone\02_system - Updated.xlsx'

# Bảng ánh xạ kết quả (system ID -> Pass/Fail)
RESULTS = {
    # Feature_RL
    'REG_01': 'Pass', 'REG_02': 'Pass', 'REG_03': 'Pass',
    'REG_05': 'Pass',
    'INT_01': 'Pass', 'INT_02': 'Pass', 'INT_03': 'N/A',
    'LOG_01': 'Pass', 'LOG_02': 'Pass', 'LOG_03': 'Pass',
    'LOG_04': 'Pass', 'LOG_05': 'N/A',
    # Feature_IM
    'IM_IC_01': 'Pass', 'IM_IC_02': 'Pass', 'IM_IC_03': 'Pass',
    'IM_IC_04': 'N/A',  'IM_IC_05': 'Pass', 'IM_IC_06': 'Pass',
    'IM_IC_07': 'Pass', 'IM_IC_08': 'N/A',  'IM_IC_09': 'N/A',
    'IM_01': 'Pass', 'IM_02': 'N/A', 'IM_03': 'N/A',
    'IM_04': 'N/A', 'IM_05': 'N/A', 'IM_06': 'N/A',
    'IM_07': 'Pass', 'IM_08': 'Pass', 'IM_09': 'N/A',
    'IM_10': 'Pass', 'IM_11': 'Pass', 'IM_12': 'N/A',
}

sys_wb = openpyxl.load_workbook(SYSTEM_FILE)

def unmerge_and_set(ws, row_idx, col_idx, value):
    """Giải phóng ô merged và gán giá trị mới"""
    # Tìm và giải phóng merged region nếu có
    merges_to_remove = []
    for merge_range in ws.merged_cells.ranges:
        if (merge_range.min_row <= row_idx <= merge_range.max_row and
                merge_range.min_col <= col_idx <= merge_range.max_col):
            merges_to_remove.append(str(merge_range))
            break
    for mr in merges_to_remove:
        ws.merged_cells.remove(mr)
    # Gán giá trị
    ws.cell(row=row_idx, column=col_idx).value = value

updated = 0
for sheet_name in ['Feature_IM', 'Feature_RL']:
    ws_sys = sys_wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    
    # Xác định cột "Testing Result" (cột AC = 29)
    result_col = 29  # Cột AC
    
    for row in ws_sys.iter_rows(min_row=10):
        test_id_cell = row[1]  # Cột B
        if test_id_cell and test_id_cell.value:
            tid = str(test_id_cell.value).strip()
            if tid in RESULTS:
                row_idx = test_id_cell.row
                # Gán giá trị vào cột result_col
                unmerge_and_set(ws_sys, row_idx, result_col, RESULTS[tid])
                print(f"  {tid} -> {RESULTS[tid]} (hàng {row_idx}, cột {result_col})")
                updated += 1

sys_wb.save(OUT_FILE)
print(f"\nĐã cập nhật {updated} dòng vào {OUT_FILE}")
